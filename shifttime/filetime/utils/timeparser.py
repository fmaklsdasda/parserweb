from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.cell import MergedCell
import re
from datetime import datetime


class ScheduleParser:
    def __init__(self, excel_file):
        self.days = []
        self.wb = load_workbook(filename=excel_file)
        self.ws = self.wb.active
        self.date_column = 1
        self.first_row = 2
        self.last_date_col = None
        self.groups_by_col = {}

    def get_merged_cell_value(self, cell):
        if isinstance(cell, MergedCell):
            for merged_range in self.ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    if merged_range.min_col == cell.column:
                        top_left_cell = self.ws.cell(
                            row=merged_range.min_row, column=merged_range.min_col
                        )
                        return top_left_cell.value
        else:
            return cell.value
        return None

    def parse_date(self, col: str):
        if col:
            date_match = re.search(r"\d{2}\.\d{2}\.\d{4}", col)
            if date_match:
                date_str = date_match.group()
                date_obj = datetime.strptime(date_str, "%d.%m.%Y")
                return date_obj
        return False

    def parse_subject(self, cell: Cell):
        val = self.get_merged_cell_value(cell)
        if val:
            teacher_pattern = r"([А-ЯЁа-яё]+ [А-ЯЁ]\.[А-ЯЁ]\.)"
            teacher_match = re.search(teacher_pattern, val)
            if teacher_match:
                subject = val[:teacher_match.start()].strip()
                teacher = re.sub(r'\s+', ' ', teacher_match.group(1).strip())
                
                room_text = val[teacher_match.end():].strip()
                room_pattern = r"(?i)(каб(?:инет)?\.?\s*\d+(?:[^\n]*)|цок)"
                room_match = re.search(room_pattern, room_text)
                room = room_match.group(0).strip() if room_match else ""
                
                return subject, teacher, room
        return False


    def parse_schedule(self):
        max_row = self.ws.max_row
        max_col = self.ws.max_column

        header_row = self.ws[2]
        for col_idx, cell in enumerate(header_row, start=1):
            value = self.get_merged_cell_value(cell)
            if value and isinstance(value, str) and re.search(r"\d", value):
                self.groups_by_col[col_idx] = value.strip()

        iter_rows = self.ws.iter_rows(
            min_row=self.first_row,
            max_row=max_row,
            min_col=self.date_column,
            max_col=max_col,
        )

        for row in iter_rows:
            date_col = row[0].value
            lesson_num = row[1].value

            dt = None

            if date_col is not None:
                dt = self.parse_date(date_col)
                if dt:
                    self.last_date_col = dt
            else:
                dt = self.last_date_col

            day = []
            if dt:
                for col_idx, col in enumerate(row[2:], start=3):
                    result = self.parse_subject(col)
                    if result:
                        subj, teacher, room = result
                        group_str = self.groups_by_col.get(col_idx, "")
                        groups = group_str.split("\n")
                        for group in groups:
                            pair = {
                                "subj": subj,
                                "group": group.strip(),
                                "dt": dt,
                                "lesson_num": lesson_num,
                                "teacher": teacher,
                                "room": room,
                            }
                            day.append(pair)

            self.days.append(day)
